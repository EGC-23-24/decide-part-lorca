from django.db.utils import IntegrityError
from django.core.exceptions import ObjectDoesNotExist
from django.shortcuts import render
from django.views.generic.base import TemplateView
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.contrib import messages
from django.urls import reverse
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from rest_framework import generics
from rest_framework.response import Response
from rest_framework.status import (
    HTTP_201_CREATED as ST_201,
    HTTP_204_NO_CONTENT as ST_204,
    HTTP_401_UNAUTHORIZED as ST_401,
    HTTP_409_CONFLICT as ST_409,
)

from base.perms import UserIsStaff
from .forms import CreationCensusForm
from .models import Census
from voting.models import Voting
from openpyxl.styles import Alignment, Font, PatternFill
from django.shortcuts import redirect
from django.core.exceptions import ValidationError
from django.contrib.auth.models import User


def census(request):
    """
    Renders the census page.

    :param request: The HTTP request object.
    :return: Rendered web page for the census.
    """
    
    return render(request, "census/census.html")

def is_admin(user):
    """
    Determines if the given user is an admin.

    :param user: The user object to be checked.
    :type user: User
    :return: True if the user is authenticated and is a staff member, False otherwise.
    :rtype: bool
    """
    
    return user.is_authenticated and user.is_staff

class CensusCreate(generics.ListCreateAPIView):
    """
    API view for listing and creating census records.

    Inherits from ListCreateAPIView to provide handling for GET and POST requests for Census model.
    """
    
    permission_classes = (UserIsStaff,)

    def create(self, request, *args, **kwargs):
        """
        Creates new census records from the provided data.

        :param request: The HTTP request object containing voting and voter data.
        :param args: Variable arguments.
        :param kwargs: Keyword arguments.
        :return: A Response object with creation status.
        """
        
        voting_id = request.data.get("voting_id")
        voters = request.data.get("voters")
        try:
            for voter in voters:
                census = Census(voting_id=voting_id, voter_id=voter)
                census.save()
        except Exception:
            return Response("Error try to create census", status=ST_409)
        return Response("Census created", status=ST_201)

    def list(self, request, *args, **kwargs):
        """
        Lists voters in a census for a given voting.

        :param request: The HTTP request object containing the voting_id parameter.
        :param args: Variable arguments.
        :param kwargs: Keyword arguments.
        :return: A Response object with the list of voters.
        """
        
        voting_id = request.GET.get("voting_id")
        voters = Census.objects.filter(voting_id=voting_id).values_list(
            "voter_id", flat=True
        )
        return Response({"voters": voters})


class CensusDetail(generics.RetrieveDestroyAPIView):
    """
    API view for retrieving and deleting census records.

    Inherits from RetrieveDestroyAPIView to provide handling for GET and DELETE requests for the Census model.
    """
    
    def destroy(self, request, voting_id, *args, **kwargs):
        """
        Deletes voters from a census based on provided voting_id and voter ids.

        :param request: The HTTP request object containing voter data.
        :param voting_id: The ID of the voting.
        :param args: Variable arguments.
        :param kwargs: Keyword arguments.
        :return: A Response object confirming deletion.
        """
        
        voters = request.data.get("voters")
        census = Census.objects.filter(voting_id=voting_id, voter_id__in=voters)
        census.delete()
        return Response("Voters deleted from census", status=ST_204)

    def retrieve(self, request, voting_id, *args, **kwargs):
        """
        Retrieves a specific voter in a census based on provided voting_id and voter id.

        :param request: The HTTP request object containing the voter_id parameter.
        :param voting_id: The ID of the voting.
        :param args: Variable arguments.
        :param kwargs: Keyword arguments.
        :return: A Response object indicating if the voter is valid.
        """

        voter = request.GET.get("voter_id")
        try:
            Census.objects.get(voting_id=voting_id, voter_id=voter)
        except ObjectDoesNotExist:
            return Response("Invalid voter", status=ST_401)
        return Response("Valid voter")


def GetId(request):
    """
    Retrieves and displays the census details for a given voting ID.

    :param request: The HTTP request object containing the voting ID.
    :return: Rendered web page with census details or an error message.
    """

    id = request.GET["id"]

    census = Census.objects.filter(voting_id=int(id))
    if len(census) == 0:
        return render(
            request,
            "census/census.html",
            {"error_id": "There is not a census with that voting_id"},
        )
    else:
        return render(request, "census/census_details.html", {"census": census})


def createCensus(request):
    """
    Handles the creation of a census record through a form submission.

    :param request: The HTTP request object.
    :return: Redirects to the census page on success or renders the creation form with errors.
    """
    
    if request.method == "POST":
        voting_id = request.POST.get("voting_id")
        voter_id = request.POST.get("voter_id")

        try:
            census = Census.objects.create(voting_id=voting_id, voter_id=voter_id)
            census.full_clean()
            census.save()
            messages.success(request, "Census created successfully")
            return redirect("census")

        # Si hay un ValidationError, muestra el mensaje de error en la página de creación del censo
        except ValidationError as e:
            if not Voting.objects.filter(id=voting_id).exists():
                return render(
                    request,
                    "census/census_create.html",
                    {
                        "error": "Voting with this ID does not exist.",
                        "form": CreationCensusForm,
                    },
                )

            if not User.objects.filter(id=voter_id).exists():
                return render(
                    request,
                    "census/census_create.html",
                    {
                        "error": "User with this ID does not exist.",
                        "form": CreationCensusForm,
                    },
                )

    # Si el método no es POST, muestra la página de creación del censo
    return render(request, "census/census_create.html", {"form": CreationCensusForm})


def deleteCensus(request):
    """
    Deletes a specific census record based on the voting and voter ID provided in the request.

    :param request: The HTTP request object containing the voting and voter IDs.
    :return: Redirects to the census page on successful deletion or renders the census page with an error message.
    """
    
    census = Census.objects.filter(
        voting_id=request.POST["voting_id"], voter_id=request.POST["voter_id"]
    )
    if len(census) == 0:
        return render(
            request,
            "census/census.html",
            {"error": "Census does not exist. Try other census"},
        )
    if len(census) != 0:
        census.delete()
        messages.success(request, "Census deleted successfully")
        return redirect("census")


def censusList(request):
    """
    Displays a list of all census records.

    :param request: The HTTP request object.
    :return: Rendered web page displaying a list of all census records.
    """
    
    queryset = Census.objects.all()
    return render(request, "census/census_list.html", {"queryset": queryset})


class CensusExportView(TemplateView):
    """
    View for exporting census data to an Excel file.

    Inherits from TemplateView to display a page for selecting the voting to export census data.
    """

    template_name = "census/export_census.html"

    def dispatch(self, request, *args, **kwargs):
        """
        Handles the request before it reaches the get or post method. Redirects to home if the user is not an admin.

        :param request: The HTTP request object.
        :param args: Variable arguments.
        :param kwargs: Keyword arguments.
        :return: HttpResponse or redirection.
        """
        
        if not is_admin(request.user):
            messages.error(request, "You must be an admin to access this page!")
            return HttpResponseRedirect("/")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        """
        Provides the context data for rendering the template.

        :param kwargs: Keyword arguments.
        :return: Context data dictionary.
        """
        
        context = super().get_context_data(**kwargs)
        votings = Voting.objects.all()
        context["votings"] = votings
        return context


def export_census(request, voting_id):
    """
    Exports the census data for a given voting to an Excel file.

    :param request: The HTTP request object.
    :param voting_id: The ID of the voting whose census data is to be exported.
    :return: HttpResponse with the Excel file.
    """

    voting = Voting.objects.get(id=voting_id)
    census = Census.objects.filter(voting_id=voting_id)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=census.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Census"

    worksheet.merge_cells("A1:B1")
    first_cell = worksheet["A1"]
    first_cell.value = "Census for: " + voting.name
    first_cell.fill = PatternFill("solid", fgColor="0D98BA")
    first_cell.font = Font(bold=True, color="FFFFFF", name="Calibri")
    first_cell.alignment = Alignment(horizontal="center")

    voting_id_title = worksheet["A2"]
    voting_id_title.value = "Voting ID"
    voting_id_title.font = Font(bold=True, color="0D98BA", name="Calibri")
    voting_id_title.alignment = Alignment(horizontal="center")

    voter_id_title = worksheet["B2"]
    voter_id_title.value = "Voter ID"
    voter_id_title.font = Font(bold=True, color="0D98BA", name="Calibri")
    voter_id_title.alignment = Alignment(horizontal="center")

    for row in census:
        row_data = [row.voting_id, row.voter_id]
        worksheet.append(row_data)

    workbook.save(response)

    return response


class CensusImportView(TemplateView):
    """
    View for importing census data from an Excel file.

    Inherits from TemplateView to display a page for selecting the voting and uploading the Excel file for census data.
    """
    
    template_name = "census/import_census.html"

    def dispatch(self, request, *args, **kwargs):
        """
        Handles the request before it reaches the get or post method. Redirects to home if the user is not an admin.

        :param request: The HTTP request object.
        :param args: Variable arguments.
        :param kwargs: Keyword arguments.
        :return: HttpResponse or redirection.
        """
        
        if not is_admin(request.user):
            messages.error(request, "You must be an admin to access this page!")
            return HttpResponseRedirect("/")
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        """
        Handles the POST request for importing census data from an uploaded Excel file.

        :param request: The HTTP request object.
        :param args: Variable arguments.
        :param kwargs: Keyword arguments.
        :return: Redirects to the import page with a success or error message.
        """
        
        voting_id = request.POST.get("voting_id")
        voting = Voting.objects.get(id=voting_id)

        if voting.end_date:
            messages.error(request, "Voting has already ended!")
            return HttpResponseRedirect(reverse("import_census"))

        if not voting.start_date:
            messages.error(request, "Voting should be started first!")
            return HttpResponseRedirect(reverse("import_census"))

        if request.method == "POST" and request.FILES:
            try:
                file = request.FILES["file"]
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    voter_id = row[0]
                    Census.objects.create(voting_id=voting_id, voter_id=voter_id)

            except Exception as e:
                messages.error(request, f"Error importing data: {str(e)}")
                return HttpResponseRedirect(reverse("import_census"))

            messages.success(request, "Data imported successfully!")
            return HttpResponseRedirect(reverse("import_census"))

        if request.method == "POST" and not request.FILES:
            messages.error(request, "No file selected!")
            return HttpResponseRedirect(reverse("import_census"))

    def get_context_data(self, **kwargs):
        """
        Provides the context data for rendering the template.

        :param kwargs: Keyword arguments.
        :return: Context data dictionary.
        """
        
        context = super().get_context_data(**kwargs)
        votings = Voting.objects.all()
        context["votings"] = votings
        return context
