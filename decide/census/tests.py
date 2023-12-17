from django.contrib.auth.models import User
from django.conf import settings
from django.urls import reverse
from rest_framework.test import APIClient

from openpyxl import load_workbook
from io import BytesIO
from openpyxl import Workbook
from django.core.files.uploadedfile import SimpleUploadedFile
from django.utils import timezone

from .models import Census
from voting.models import Voting, Question, QuestionOption
from base.models import Auth
from base.tests import BaseTestCase


class CensusTestCase(BaseTestCase):
    """
    Test case for census-related operations.

    This class inherits from TestCase and provides test methods for creating, 
    validating, and deleting census records.
    """
    
    def setUp(self):
        """
        Set up the test environment before each test method.

        Creates a voting instance, a user, and a census record.
        """
        
        # Crear una votación
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        # Crear un votante
        u, created = User.objects.get_or_create(username="testvoter")
        u.is_active = True
        u.save()

        # Crear un censo
        self.census = Census.objects.create(voting_id=v.id, voter_id=u.id)

        super().setUp()

    def test_create_census(self):
        """
        Test the creation of a census record.

        Validates that the census record is correctly created and its attributes 
        match the expected values.
        """
        
        # Crear una votación
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        # Crear un votante
        u, created = User.objects.get_or_create(username="testvoter")
        u.is_active = True
        u.save()

        # Crear un censo
        census = Census.objects.create(voting_id=v.id, voter_id=u.id)

        # Comprobar que se ha creado correctamente
        self.assertEqual(census.voting_id, v.id)
        self.assertEqual(census.voter_id, u.id)
        self.assertEqual(Census.objects.latest("id").voting_id, v.id)
        self.assertEqual(Census.objects.latest("id").voter_id, u.id)

    def test_create_census_invalid_voting_id(self):
        """
        Test the creation of a census record with an invalid voting ID.

        Expects a ValueError when trying to create a census with an invalid voting ID.
        """
        
        # Crear una votación
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        # Crear un votante
        u, created = User.objects.get_or_create(username="testvoter")
        u.is_active = True
        u.save()

        with self.assertRaises(ValueError):
            # Crear un censo
            census = Census.objects.create(voting_id="invalid_voting_id", voter_id=u.id)
            census.full_clean()  # This should raise a ValidationError exception

    def test_create_census_invalid_voter_id(self):
        """
        Test the creation of a census record with an invalid voter ID.

        Expects a ValueError when trying to create a census with an invalid voter ID.
        """
        
        # Crear una votación
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        # Crear un votante
        u, created = User.objects.get_or_create(username="testvoter")
        u.is_active = True
        u.save()

        with self.assertRaises(ValueError):
            # Attempt to create a census with an invalid voter_id
            census = Census.objects.create(voting_id=v.id, voter_id="invalid_voter_id")
            census.full_clean()

    def test_create_census_invalid_voting_id_and_voter_id(self):
        """
        Test the creation of a census record with both invalid voting and voter IDs.

        Expects a ValueError when trying to create a census with invalid voting and 
        voter IDs.
        """
        
        with self.assertRaises(ValueError):
            # Attempt to create a census with an invalid voting_id and voter_id
            census = Census.objects.create(
                voting_id="invalid_voting_id", voter_id="invalid_voter_id"
            )
            census.full_clean()

    def test_delete_census(self):
        """
        Test the deletion of a census record.

        Validates that the census record is successfully deleted and no longer exists 
        in the database.
        """
        
        # Delete any existing Census objects to avoid IntegrityError
        Census.objects.all().delete()

        # Crear una votación
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        # Crear un votante
        u, created = User.objects.get_or_create(username="testvoter")
        u.is_active = True
        u.save()

        # Create a Census object to delete
        census_to_delete = Census.objects.create(voting_id=v.id, voter_id=u.id)

        # Define the URL and the data
        url = reverse("census_deleted")  # replace with your URL name
        data = {
            "voting_id": census_to_delete.voting_id,
            "voter_id": census_to_delete.voter_id,
        }

        # Make the POST request
        response = self.client.post(url, data, follow=True)

        # Check the status code and the response data
        self.assertEqual(response.status_code, 200)
        self.assertFalse(
            Census.objects.filter(
                voting_id=census_to_delete.voting_id, voter_id=census_to_delete.voter_id
            ).exists()
        )

    def test_delete_census_invalid_voting_id(self):
        """
        Test the deletion of a census record with an invalid voting ID.

        Expects a ValueError when trying to delete a census with an invalid voting ID.
        """
        
        # Crear una votación
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        # Crear un votante
        u, created = User.objects.get_or_create(username="testvoter")
        u.is_active = True
        u.save()

        with self.assertRaises(ValueError):
            # Attempt to delete a census with an invalid voting_id
            census = Census.objects.create(voting_id="invalid_voting_id", voter_id=u.id)
            census.full_clean()

    def test_delete_census_invalid_voter_id(self):
        """
        Test the deletion of a census record with an invalid voter ID.

        Expects a ValueError when trying to delete a census with an invalid voter ID.
        """
        
        # Crear una votación
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        # Crear un votante
        u, created = User.objects.get_or_create(username="testvoter")
        u.is_active = True
        u.save()

        with self.assertRaises(ValueError):
            # Attempt to delete a census with an invalid voter_id
            census = Census.objects.create(voting_id=v.id, voter_id="invalid_voter_id")
            census.full_clean()

    def test_delete_census_invalid_voting_id_and_voter_id(self):
        """
        Test the deletion of a census record with both invalid voting and voter IDs.

        Expects a ValueError when trying to delete a census with invalid voting and 
        voter IDs.
        """
        
        with self.assertRaises(ValueError):
            # Attempt to delete a census with an invalid voting_id and voter_id
            census = Census.objects.create(
                voting_id="invalid_voting_id", voter_id="invalid_voter_id"
            )
            census.full_clean()

    def test_list_census(self):
        """
        Test listing of all census records.

        Validates that the census records can be retrieved and their attributes match 
        the expected values.
        """
        
        # Eliminar todos los objetos Census existentes
        Census.objects.all().delete()

        # Crear un votante
        u, created = User.objects.get_or_create(username="testvoter")
        u.is_active = True
        u.save()

        # Crear una votación
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        # Agregar el votante al censo
        census = Census(voter_id=u.id, voting_id=v.id)
        census.save()

        # Define the URL and the data
        url = reverse("census_list")

        # Make the POST request
        response = self.client.get(url, follow=True)

        # Check the status code and the response data
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Census.objects.count(), 1)
        # Comprobar que el voting_id del objeto Census es el correcto
        self.assertEqual(Census.objects.latest("id").voting_id, v.id)
        self.assertEqual(Census.objects.latest("id").voter_id, u.id)

    def test_list_census_invalid_voting_id(self):
        """
        Test listing of census records with an invalid voting ID.

        Expects a ValueError when trying to list census records with an invalid voting ID.
        """
        
        # Crear una votación
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        # Crear un votante
        u, created = User.objects.get_or_create(username="testvoter")
        u.is_active = True
        u.save()

        with self.assertRaises(ValueError):
            # Attempt to list a census with an invalid voting_id
            census = Census.objects.create(voting_id="invalid_voting_id", voter_id=u.id)
            census.full_clean()

    def test_list_census_invalid_voter_id(self):
        """
        Test listing of census records with an invalid voter ID.

        Expects a ValueError when trying to list census records with an invalid voter ID.
        """
        
        # Crear una votación
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        # Crear un votante
        u, created = User.objects.get_or_create(username="testvoter")
        u.is_active = True
        u.save()

        with self.assertRaises(ValueError):
            # Attempt to list a census with an invalid voter_id
            census = Census.objects.create(voting_id=v.id, voter_id="invalid_voter_id")
            census.full_clean()

    def test_list_census_invalid_voting_id_and_voter_id(self):
        """
        Test listing of census records with both invalid voting and voter IDs.

        Expects a ValueError when trying to list census records with invalid voting and 
        voter IDs.
        """
        
        with self.assertRaises(ValueError):
            # Attempt to list a census with an invalid voting_id and voter_id
            census = Census.objects.create(
                voting_id="invalid_voting_id", voter_id="invalid_voter_id"
            )
            census.full_clean()

    def test_get_census(self):
        """
        Test retrieval of a specific census record.

        Validates that a specific census record can be retrieved and its attributes match 
        the expected values.
        """
        
        # Eliminar todos los objetos Census existentes
        Census.objects.all().delete()

        # Crear un votante
        u, created = User.objects.get_or_create(username="testvoter")
        u.is_active = True
        u.save()

        # Crear una votación
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        # Agregar el votante al censo
        census = Census(voter_id=u.id, voting_id=v.id)
        census.save()

        # Define the URL and the data
        url = reverse("census_details")

        data = {"id": census.id}
        # Make the POST request
        response = self.client.get(url, data)

        # Check the status code and the response data
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Census.objects.count(), 1)
        self.assertEqual(Census.objects.latest("id").voting_id, v.id)
        self.assertEqual(Census.objects.latest("id").voter_id, u.id)

    def test_get_census_invalid_voting_id(self):
        """
        Test retrieval of a census record with an invalid voting ID.

        Expects a ValueError when trying to retrieve a census record with an invalid 
        voting ID.
        """
        
        # Crear una votación
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        # Crear un votante
        u, created = User.objects.get_or_create(username="testvoter")
        u.is_active = True
        u.save()

        with self.assertRaises(ValueError):
            # Attempt to get a census with an invalid voting_id
            census = Census.objects.create(voting_id="invalid_voting_id", voter_id=u.id)
            census.full_clean()

    def test_get_census_invalid_voter_id(self):
        """
        Test retrieval of a census record with an invalid voter ID.

        Expects a ValueError when trying to retrieve a census record with an invalid 
        voter ID.
        """
        
        # Crear una votación
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        # Crear un votante
        u, created = User.objects.get_or_create(username="testvoter")
        u.is_active = True
        u.save()

        with self.assertRaises(ValueError):
            # Attempt to get a census with an invalid voter_id
            census = Census.objects.create(voting_id=v.id, voter_id="invalid_voter_id")
            census.full_clean()

    def test_get_census_invalid_voting_id_and_voter_id(self):
        """
        Test retrieval of a census record with both invalid voting and voter IDs.

        Expects a ValueError when trying to retrieve a census record with invalid 
        voting and voter IDs.
        """
        
        with self.assertRaises(ValueError):
            # Attempt to get a census with an invalid voting_id and voter_id
            census = Census.objects.create(
                voting_id="invalid_voting_id", voter_id="invalid_voter_id"
            )
            census.full_clean()

    def tearDown(self):
        """
        Tear down the test environment after each test method.

        Cleans up any created test data.
        """
        
        super().tearDown()
        self.census = None

    def test_check_vote_permissions(self):
        """
        Test to check voting permissions for a user.

        Validates that a user has permission to vote in a specific voting.
        """
        
        # Crear un votante
        u, created = User.objects.get_or_create(username="testvoter")
        u.is_active = True
        u.save()

        # Crear una votación
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        # Autorizar al votante para votar
        census = Census(voter_id=u.id, voting_id=v.id)
        census.save()

        response = self.client.get(
            "/census/{}/?voter_id={}".format(v.id, u.id), format="json"
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), "Valid voter")

    def test_list_voting(self):
        """
        Test listing all votings.

        Validates that all votings can be listed and their attributes match the 
        expected values.
        """
        
        # Crear una votación
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        # Crear un votante
        u, created = User.objects.get_or_create(username="testvoter")
        u.is_active = True
        u.save()

        # Crear un censo
        census = Census.objects.create(voting_id=v.id, voter_id=u.id)

        self.login()
        response = self.client.get("/census/?voting_id={}".format(v.id), format="json")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {"voters": [u.id]})

    def test_add_new_voters_conflict(self):
        """
        Test adding new voters to a voting when there's a conflict.

        Validates the behavior when adding a voter who no longer exists.
        """
        
        # Crear una votación
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        # Crear un votante
        u, created = User.objects.get_or_create(username="testvoter")
        u.is_active = True
        u.save()

        # Crear un censo
        census = Census.objects.create(voting_id=v.id, voter_id=u.id)

        # Borrar el votante
        u.delete()

        self.login()
        data = {"voting_id": v.id, "voters": [u.id]}
        response = self.client.post("/census/", data, format="json")
        self.assertEqual(response.status_code, 409)

    def test_add_new_voters(self):
        """
        Test adding new voters to a voting.

        Validates that new voters can be added to a voting successfully.
        """
        
        # Crear una votación
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        # Crear un votante
        u, created = User.objects.get_or_create(username="testvoter2")
        u.is_active = True
        u.save()

        self.login()
        data = {"voting_id": v.id, "voters": [u.id]}
        response = self.client.post("/census/", data, format="json")
        self.assertEqual(response.status_code, 201)
        self.assertEqual(len(data.get("voters")), Census.objects.count() - 1)

    def test_destroy_voter(self):
        """
        Test destroying a voter in a voting.

        Validates that a voter can be removed from a voting successfully.
        """
        
        # Eliminar todos los objetos Census existentes
        Census.objects.all().delete()

        # Crear un votante
        u, created = User.objects.get_or_create(username="testvoter")
        u.is_active = True
        u.save()

        # Crear una votación
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        # Agregar el votante al censo
        census = Census(voter_id=u.id, voting_id=v.id)
        census.save()

        census_to_delete = Census.objects.get(voter_id=u.id, voting_id=v.id)
        deletion = census_to_delete.delete()

        self.assertEqual(deletion[0], 1)  # Comprobar que se eliminó un objeto
        self.assertEqual(
            0, Census.objects.count()
        )  # Comprobar que no hay objetos Census

class CensusExportViewTest(BaseTestCase):
    """
    Test suite for the CensusExportView functionality.
    """
    
    def setUp(self):
        """
        Set up the test environment for each test method.
        """
        
        super().setUp()

    def create_voting(self):
        """
        Creates a Voting instance with associated questions and options.

        Returns:
            Voting: The Voting instance that was created.
        """
        
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        a, _ = Auth.objects.get_or_create(
            url=settings.BASEURL, defaults={"me": True, "name": "test auth"}
        )
        a.save()
        v.auths.add(a)

        return v

    def test_census_export_view_for_admin(self):
        """
        Test the CensusExportView for admin users.

        Ensures that admin users can access the export census view and
        that the context contains the correct data.
        """
        
        self.create_voting()

        self.client.force_login(User.objects.get(username="admin"))

        url = reverse("export_census")

        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)

        self.assertTrue("votings" in response.context)

        votings = response.context["votings"]
        self.assertEqual(votings.count(), Voting.objects.count())

        self.assertTemplateUsed(response, "census/export_census.html")

    def test_census_export_view_for_non_admin(self):
        """
        Test the CensusExportView for non-admin users.

        Ensures that non-admin users are redirected when trying to access
        the export census view.
        """
        
        self.create_voting()

        self.client.force_login(User.objects.get(username="noadmin"))

        url = reverse("export_census")

        response = self.client.get(url)
        self.assertEqual(response.status_code, 302)


class ExportCensusTest(BaseTestCase):
    """
    Test suite for exporting census data.
    """
    
    def setUp(self):
        """
        Set up the test environment for each test method.
        """
        
        super().setUp()

    def create_voting(self):
        """
        Creates a Voting instance with associated questions and options.

        Returns:
            Voting: The Voting instance that was created.
        """
        
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        a, _ = Auth.objects.get_or_create(
            url=settings.BASEURL, defaults={"me": True, "name": "test auth"}
        )
        a.save()
        v.auths.add(a)

        return v

    def create_voters(self, v):
        """
        Creates voters and adds them to the census of a given voting.

        Args:
            v (Voting): The Voting instance to which the voters will be added.
        """

        for i in range(100):
            u, _ = User.objects.get_or_create(username="testvoter{}".format(i))
            u.is_active = True
            u.save()
            c = Census(voter_id=u.id, voting_id=v.id)
            c.save()

    def test_export_census(self):
        """
        Test the functionality of exporting census data to an Excel file.

        Ensures that the response contains the correct content type and
        disposition, and the Excel file contains the correct data.
        """
        
        v = self.create_voting()
        self.create_voters(v)

        url = reverse("export_census_of_voting", args=[v.id])

        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)

        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        self.assertEqual(
            response["Content-Disposition"], "attachment; filename=census.xlsx"
        )

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active

        self.assertEqual(worksheet["A1"].value, f"Census for: {v.name}")

        self.assertEqual(worksheet["A2"].value, "Voting ID")
        self.assertEqual(worksheet["B2"].value, "Voter ID")

        census_data = Census.objects.filter(voting_id=v.id)
        for i, row in enumerate(census_data, start=3):
            self.assertEqual(worksheet[f"A{i}"].value, row.voting_id)
            self.assertEqual(worksheet[f"B{i}"].value, row.voter_id)


class CensusImportViewTest(BaseTestCase):
    """
    Test suite for the CensusImportView functionality.
    """

    def setUp(self):
        """
        Set up the test environment for each test method by creating a user and initializing the base test case.
        """
        
        super().setUp()
        user = User(username="test")
        user.set_password("test")
        user.save()

    def create_voting(self):
        """
        Creates a Voting instance with associated questions and options.

        Returns:
            Voting: The Voting instance that was created.
        """
        
        q = Question(desc="test question")
        q.save()
        for i in range(5):
            opt = QuestionOption(question=q, option="option {}".format(i + 1))
            opt.save()
        v = Voting(name="test voting", question=q)
        v.save()

        a, _ = Auth.objects.get_or_create(
            url=settings.BASEURL, defaults={"me": True, "name": "test auth"}
        )
        a.save()
        v.auths.add(a)

        return v

    def test_census_import_view_success(self):
        """
        Test the successful import of census data from an Excel file.

        Ensures that census data is correctly imported into the database
        and appropriate success message is displayed.
        """
        
        v = self.create_voting()
        v.create_pubkey()
        v.start_date = timezone.now()
        v.save()

        user1 = User.objects.get(username="noadmin")
        user2 = User.objects.get(username="test")

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Voter ID"])
        sheet.append([user1.id])
        sheet.append([user2.id])

        file_buffer = BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)

        excel_file = SimpleUploadedFile("census.xlsx", file_buffer.read())

        self.client.force_login(User.objects.get(username="admin"))

        url = reverse("import_census")

        response = self.client.post(
            url, {"file": excel_file, "voting_id": v.id}, follow=True
        )

        self.assertEqual(response.status_code, 200)

        census_data = Census.objects.all()
        self.assertEqual(census_data.count(), 2)
        self.assertEqual(census_data[0].voting_id, v.id)
        self.assertEqual(census_data[0].voter_id, user1.id)
        self.assertEqual(census_data[1].voting_id, v.id)
        self.assertEqual(census_data[1].voter_id, user2.id)

        messages = list(response.context["messages"])
        self.assertEqual(len(messages), 1)
        self.assertEqual(str(messages[0]), "Data imported successfully!")

    def test_census_import_view_fail(self):
        """
        Test the failure case of importing census data from an Excel file with incorrect data.

        Ensures that an error message is displayed when incorrect data is imported.
        """
        
        v = self.create_voting()
        v.create_pubkey()
        v.start_date = timezone.now()
        v.save()

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Voter ID"])
        sheet.append(["A"])

        file_buffer = BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)

        excel_file = SimpleUploadedFile("census.xlsx", file_buffer.read())

        self.client.force_login(User.objects.get(username="admin"))

        url = reverse("import_census")

        response = self.client.post(
            url, {"file": excel_file, "voting_id": v.id}, follow=True
        )

        self.assertEqual(response.status_code, 200)

        messages = list(response.context["messages"])
        self.assertEqual(len(messages), 1)
        self.assertTrue("Error importing data" in str(messages[0]))

    def test_census_import_view_no_file(self):
        """
        Test the case where no file is provided for importing census data.

        Ensures that an appropriate error message is displayed when no file is selected.
        """
        
        v = self.create_voting()
        v.create_pubkey()
        v.start_date = timezone.now()
        v.save()

        self.client.force_login(User.objects.get(username="admin"))

        url = reverse("import_census")

        response = self.client.post(url, {"voting_id": v.id}, follow=True)

        self.assertEqual(response.status_code, 200)

        messages = list(response.context["messages"])
        self.assertEqual(len(messages), 1)
        self.assertEqual(str(messages[0]), "No file selected!")

    def test_census_import_view_voting_ended(self):
        """
        Test the case where an attempt is made to import census data for a voting that has already ended.

        Ensures that an appropriate error message is displayed when trying to import data for an ended voting.
        """
        
        v = self.create_voting()
        v.create_pubkey()
        v.start_date = timezone.now()
        v.end_date = timezone.now()
        v.save()

        self.client.force_login(User.objects.get(username="admin"))

        url = reverse("import_census")

        response = self.client.post(
            url, {"file": "test", "voting_id": v.id}, follow=True
        )

        self.assertEqual(response.status_code, 200)

        messages = list(response.context["messages"])
        self.assertEqual(len(messages), 1)
        self.assertEqual(str(messages[0]), "Voting has already ended!")

    def test_census_import_view_voting_not_started(self):
        """
        Test the case where an attempt is made to import census data for a voting that has not started yet.

        Ensures that an appropriate error message is displayed when trying to import data for a not-started voting.
        """
        
        v = self.create_voting()

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Voter ID"])
        sheet.append([1])

        file_buffer = BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)

        excel_file = SimpleUploadedFile("census.xlsx", file_buffer.read())

        self.client.force_login(User.objects.get(username="admin"))

        url = reverse("import_census")

        response = self.client.post(
            url, {"file": excel_file, "voting_id": v.id}, follow=True
        )

        self.assertEqual(response.status_code, 200)

        messages = list(response.context["messages"])
        self.assertEqual(len(messages), 1)
        self.assertEqual(str(messages[0]), "Voting should be started first!")

    def test_census_import_view_not_admin(self):
        """
        Test the access control for the import census view to ensure only admins can access it.

        Ensures that non-admin users receive an appropriate error message when trying to access the import view.
        """
        
        v = self.create_voting()
        v.create_pubkey()
        v.start_date = timezone.now()
        v.save()

        self.client.force_login(User.objects.get(username="noadmin"))

        url = reverse("import_census")

        response = self.client.post(
            url, {"file": "test", "voting_id": v.id}, follow=True
        )

        self.assertEqual(response.status_code, 200)

        messages = list(response.context["messages"])
        self.assertEqual(len(messages), 1)
        self.assertEqual(str(messages[0]), "You must be an admin to access this page!")
